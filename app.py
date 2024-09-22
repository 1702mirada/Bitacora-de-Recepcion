from flask import Flask, request, jsonify
from openpyxl import Workbook, load_workbook
import os

# Crear la aplicación Flask
app = Flask(__name__)

# Ruta para guardar las entradas de la bitácora
@app.route('/save', methods=['POST'])
def save_log():
    date = request.form['date']
    time = request.form['time']
    description = request.form['description']
    pdf = request.files['pdf']
    
    file_path = 'bitacora.xlsx'
    pdf_path = os.path.join('pdfs', pdf.filename)
    
    # Crear el archivo de Excel si no existe
    if not os.path.exists(file_path):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(['Fecha', 'Hora', 'Descripción', 'PDF'])
        workbook.save(file_path)
    
    # Guardar el archivo PDF
    if not os.path.exists('pdfs'):
        os.makedirs('pdfs')
    pdf.save(pdf_path)
    
    # Cargar el archivo de Excel y agregar la nueva entrada
    workbook = load_workbook(file_path)
    sheet = workbook.active
    sheet.append([date, time, description, pdf_path])
    workbook.save(file_path)
    
    return jsonify({'message': 'Entrada guardada exitosamente'})

# Ruta para guardar el inventario
@app.route('/inventory', methods=['POST'])
def save_inventory():
    data = request.get_json()
    item = data['item']
    quantity = data['quantity']
    
    file_path = 'inventario.xlsx'
    
    # Crear el archivo de Excel si no existe
    if not os.path.exists(file_path):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(['Artículo', 'Cantidad'])
        workbook.save(file_path)
    
    # Cargar el archivo de Excel y agregar la nueva entrada
    workbook = load_workbook(file_path)
    sheet = workbook.active
    sheet.append([item, quantity])
    workbook.save(file_path)
    
    return jsonify({'message': 'Inventario actualizado exitosamente'})

# Ejecutar la aplicación Flask
if __name__ == '__main__':
    app.run(debug=True)
